
from bs4 import BeautifulSoup

file = "index.txt"
f = open(file, "r", encoding = "utf_8")

data = f.read()

f.close()

sp = BeautifulSoup(data, 'html.parser')

# テスト
selector =  '#yads_4670651-41'
print(sp.select_one(selector).string)
today = sp.select_one(selector).string


selector =  '#globalDate > p > strong'
print(sp.select_one(selector).string)
today = sp.select_one(selector).string

selector =  '#trPortfolioListRow1 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
nikkei = sp.select_one(selector).string

selector =  '#trPortfolioListRow2 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
topix = sp.select_one(selector).string

selector =  '#trPortfolioListRow3 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
usd = sp.select_one(selector).string

selector =  '#trPortfolioListRow4 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
usbond = sp.select_one(selector).string

selector =  '#trPortfolioListRow5 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
eur = sp.select_one(selector).string

selector =  '#trPortfolioListRow6 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
aud = sp.select_one(selector).string

selector =  '#trPortfolioListRow7 > td:nth-child(7) > strong'
print(sp.select_one(selector).string)
tur = sp.select_one(selector).string


#指標のexcelへの書き込み

import openpyxl as op
wb = op.load_workbook('pandas_資産残高.xlsx')
sheet1 = wb['指標']

sw_out = 0
cnt_retu = 1

while sw_out == 0:
    if sheet1.cell(row=1,column=cnt_retu).value == None:
        sw_out = 1
    else:
        cnt_retu += 1

cnt_retu = 2

sheet1.cell(row=1,column=cnt_retu).value = today
sheet1.cell(row=2,column=cnt_retu).value = nikkei
sheet1.cell(row=3,column=cnt_retu).value = topix
sheet1.cell(row=4,column=cnt_retu).value = usd
sheet1.cell(row=5,column=cnt_retu).value = usbond
sheet1.cell(row=6,column=cnt_retu).value = eur
sheet1.cell(row=7,column=cnt_retu).value = aud
sheet1.cell(row=8,column=cnt_retu).value = tur

wb.save('pandas_資産残高.xlsx')
