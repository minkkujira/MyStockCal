
from bs4 import BeautifulSoup
import pandas as pd
import sys
import openpyxl as op

df_zokusei = pd.read_excel(
    'pandas_商品属性.xlsx', sheet_name='商品属性', header=0, index_col=0)
# df_yokin = pd.read_excel('pandas_資産残高.xlsx',sheet_name='預金',
# header=0,index_col=None)
# df_kabu = pd.read_excel('pandas_資産残高.xlsx',sheet_name='株式',
# header=0,index_col=None)
# df_toshi


# 新規商品属性追加処理
def zokusei_update(df_work, add_syohin):
    wk1_dic = {1:'株', 2:'債券', 3:'預金',4:'定期',5:'定期'} 
    wk2_dic = {1:'日本', 2:'米国', 3:'ユーロ',4:'豪州',5:'マルチ'}
    wk3_dic = {1:'投信', 2:''}
    wk4_dic = {1:'SB', 2:''}
    
    while True:
        print('商品名は', add_syohin)
        print('種類は？1：株　2：債券　3：預金　4：定期　5：定期')
        wk1 = int(input())
        if wk1 in wk1_dic:
            print('国は？1：日本　2：米国　3：ユーロ　4：豪州　5：マルチ')
            wk2 = int(input())
            if wk2 in wk2_dic:
                print('形態は？1：投信　2：その他')
                wk3 = int(input())
                if wk3 in wk3_dic:
                    print('ソフトバンク？1：ＳＢ　2：その他')
                    wk4 = int(input())
                    if wk4 in wk4_dic:
                        print('登録終了') 
                        wk5 = None
                        df_work.loc[add_syohin] = [wk1_dic[wk1], wk2_dic[wk2], wk3_dic[wk3], wk5, wk5, wk4_dic[wk4]]
                        # print(df_work)
                        break
                    else:
                        print('入力エラー')
                else:
                    print('入力エラー')
            else:
                print('入力エラー')
        else:
            print('入力エラー')


file = "port.text"
f = open(file, "r", encoding="utf_8")

data = f.read()

f.close()

sp = BeautifulSoup(data, 'html.parser')


# 預金現金仮想通貨
selector = "#portfolio_det_depo > section > table > tbody td"

n = len(sp.select(selector))

for count in range(n):
    m = count % 5
    if m == 0:
        syohin = sp.select(selector)[count].string
        sw_hantei = syohin in df_zokusei.index
        if sw_hantei == False:
            # print('預金  ：', syohin)
            # print('属性無し処理')
            zokusei_update(df_zokusei, syohin)
            df_zokusei.to_excel('pandas_商品属性.xlsx', sheet_name='商品属性')
            # sys.exit()

# 株式（現物）
selector = "#portfolio_det_eq > table > tbody td"

n = len(sp.select(selector))

for count in range(n):
    m = count % 13
    if m == 1:
        syohin = sp.select(selector)[count].string
        sw_hantei = syohin in df_zokusei.index
        if sw_hantei == False:
            # print('株式  ：', syohin)
            # print('属性無し処理')
            zokusei_update(df_zokusei, syohin)
            df_zokusei.to_excel('pandas_商品属性.xlsx', sheet_name='商品属性')
            # sys.exit()


# 投資信託
selector = "#portfolio_det_mf > table > tbody td"

n = len(sp.select(selector))

for count in range(n):
    m = count % 12
    if m == 0:
        syohin = sp.select(selector)[count].string
        sw_hantei = syohin in df_zokusei.index
        if sw_hantei == False:
            # print('投資信託  ：', syohin)
            # print('属性無し処理')
            zokusei_update(df_zokusei, syohin)
            df_zokusei.to_excel('pandas_商品属性.xlsx', sheet_name='商品属性')
            # sys.exit()


# 債券
selector = "#portfolio_det_bd > table > tbody td"

n = len(sp.select(selector))

for count in range(n):
    m = count % 5
    if m == 0:
        syohin = sp.select(selector)[count].string
        sw_hantei = syohin in df_zokusei.index
        if sw_hantei == False:
            # print(count)
            # print('債券  ：', syohin)
            # print('属性無し処理')
            zokusei_update(df_zokusei, syohin)
            df_zokusei.to_excel('pandas_商品属性.xlsx', sheet_name='商品属性')
#            sys.exit()


# その他の資産
selector = "#portfolio_det_oth > table > tbody td"

n = len(sp.select(selector))

for count in range(n):
    m = count % 9
    if m == 0:
        syohin = sp.select(selector)[count].string
        sw_hantei = syohin in df_zokusei.index
        if sw_hantei == False:
            # print(count)
            # print('その他資産  ：', syohin)
            # print('属性無し処理')
            zokusei_update(df_zokusei, syohin)
            df_zokusei.to_excel('pandas_商品属性.xlsx', sheet_name='商品属性')
            # sys.exit()


# 残高書き込み

wb = op.load_workbook('pandas_資産残高.xlsx')

# 1預金現金仮想通貨

sheet1 = wb['預金']
# 残高クリア
for cnt_row in range(2, sheet1.max_row+1):
    for cnt_col in range(1, sheet1.max_column):
        sheet1.cell(row=cnt_row, column=cnt_col).value = None


selector = "#portfolio_det_depo > section > table > tbody td"

n = len(sp.select(selector))
m1 = 2
for count in range(n):
    m = count % 5
    if m == 0:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=1).value = syohin
    if m == 1:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=2).value = syohin
        m1 = m1 + 1

# 1続けて債券

selector = "#portfolio_det_bd > table > tbody td"

n = len(sp.select(selector))
for count in range(n):
    m = count % 5
    if m == 0:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=1).value = syohin
    if m == 1:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=2).value = syohin
        m1 = m1 + 1


# 2株式

sheet1 = wb['株式']

# 残高クリア
for cnt_row in range(2, sheet1.max_row+1):
    for cnt_col in range(1, sheet1.max_column):
        sheet1.cell(row=cnt_row, column=cnt_col).value = None

selector = "#portfolio_det_eq > table > tbody td"

n = len(sp.select(selector))
m1 = 2
for count in range(n):
    m = count % 13
    if m == 1:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=2).value = syohin
    if m == 5:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=6).value = syohin
        m1 = m1 + 1

# 3投資信託

sheet1 = wb['投信']

# 残高クリア
for cnt_row in range(2, sheet1.max_row+1):
    for cnt_col in range(1, sheet1.max_column):
        sheet1.cell(row=cnt_row, column=cnt_col).value = None

selector = "#portfolio_det_mf > table > tbody td"

n = len(sp.select(selector))
m1 = 2
for count in range(n):
    m = count % 12
    if m == 0:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=1).value = syohin
    if m == 4:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=5).value = syohin
        m1 = m1 + 1

# 4続けてその他の資産

selector = "#portfolio_det_oth > table > tbody td"

n = len(sp.select(selector))

for count in range(n):
    m = count % 9
    if m == 0:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=1).value = syohin
    if m == 2:
        syohin = sp.select(selector)[count].string
        sheet1.cell(row=m1, column=5).value = syohin
        m1 = m1 + 1

wb.save('pandas_資産残高.xlsx')
