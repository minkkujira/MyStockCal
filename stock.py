import pandas as pd

df_zokusei = pd.read_excel('pandas_商品属性.xlsx', sheet_name='商品属性', header=0,
                           index_col=0)
df_yokin = pd.read_excel('pandas_資産残高.xlsx', sheet_name='預金', header=0,
                         index_col=None)
df_kabu = pd.read_excel('pandas_資産残高.xlsx', sheet_name='株式', header=0,
                        index_col=None)
df_toshin = pd.read_excel('pandas_資産残高.xlsx', sheet_name='投信', header=0,
                          index_col=None)

zan_all = 0
zan_yokin = 0
zan_teiki = 0
zan_kabu = 0
zan_saiken = 0
zan_reit = 0
zan_yen = 0
zan_usa = 0
zan_eur = 0
zan_aus = 0
zan_out = 0
zan_sb = 0

item_add1 = 0
item_add2 = 0
item_add3 = 0
item_add4 = 0


def choice():
    global item_add1
    global item_add2
    global item_add3
    global item_add4
    
    ch_1 = val_1.get()
    print(ch_1)
    print('種類は　' + item_1[ch_1])
    item_add1 = item_1[ch_1]
    print('読込時add1', item_add1)

    ch_2 = val_2.get()
    print(ch_2)
    print('国は　' + item_2[ch_2])
    item_add2 = item_2[ch_2]

    ch_3 = val_3.get()
    print(ch_3)
    print('形態は　' + item_3[ch_3])
    item_add3 = item_3[ch_3]

    ch_4 = val_4.get()
    print(ch_4)
    print('ＳＢは　' + item_4[ch_4])
    item_add4 = item_4[ch_4]

# 預金債券計算

for row in df_yokin.values:
    syohin = row[0]
    # print('商品', syohin)

# 円を削除
    zan_a = row[1]
    zan_a = zan_a[:-1]
    zan_a = int(zan_a.replace(',', ''))
    zan_all += zan_a

    if df_zokusei.at[syohin, '種類'] == '預金':
        zan_yokin += zan_a
    if df_zokusei.at[syohin, '種類'] == '定期':
        zan_teiki += zan_a
    if df_zokusei.at[syohin, '種類'] == '株':
        zan_kabu += zan_a
    if df_zokusei.at[syohin, '種類'] == '債券':
        zan_saiken += zan_a
    if df_zokusei.at[syohin, '種類'] == 'リート':
        zan_reit += zan_a
    if df_zokusei.at[syohin, '国'] == '日本':
        zan_yen += zan_a
    if df_zokusei.at[syohin, '国'] == '米国':
        zan_usa += zan_a
    if df_zokusei.at[syohin, '国'] == '豪州':
        zan_aus += zan_a
    if df_zokusei.at[syohin, '国'] == 'ユーロ':
        zan_eur += zan_a
    if df_zokusei.at[syohin, '国'] == 'マルチ':
        zan_out += zan_a
    if df_zokusei.at[syohin, 'ＳＢ'] == 'SB':
        zan_sb += zan_a

# 株式計算

for row in df_kabu.values:
    syohin = row[1]

# 円を削除
    zan_a = row[5]
    zan_a = zan_a[:-1]
    zan_a = int(zan_a.replace(',', ''))
    zan_all += zan_a

    if df_zokusei.at[syohin, '種類'] == '預金':
        zan_yokin += zan_a
    if df_zokusei.at[syohin, '種類'] == '定期':
        zan_teiki += zan_a
    if df_zokusei.at[syohin, '種類'] == '株':
        zan_kabu += zan_a
    if df_zokusei.at[syohin, '種類'] == '債券':
        zan_saiken += zan_a
    if df_zokusei.at[syohin, '種類'] == 'リート':
        zan_reit += zan_a
    if df_zokusei.at[syohin, '国'] == '日本':
        zan_yen += zan_a
    if df_zokusei.at[syohin, '国'] == '米国':
        zan_usa += zan_a
    if df_zokusei.at[syohin, '国'] == '豪州':
        zan_aus += zan_a
    if df_zokusei.at[syohin, '国'] == 'ユーロ':
        zan_eur += zan_a
    if df_zokusei.at[syohin, '国'] == 'マルチ':
        zan_out += zan_a
    if df_zokusei.at[syohin, 'ＳＢ'] == 'SB':
        zan_sb += zan_a

# 投信計算

import numpy as np
import math

for row in df_toshin.values:
    syohin = row[0]

# 円を削除
    zan_a = row[4]
    
    if zan_a != zan_a:
        zan_a = '0円'

    zan_a = zan_a[:-1]
    zan_a = int(zan_a.replace(',', ''))
    zan_all += zan_a

    # print(syohin)
    if df_zokusei.at[syohin, '種類'] == '預金':
        zan_yokin += zan_a
    if df_zokusei.at[syohin, '種類'] == '定期':
        zan_teiki += zan_a
    if df_zokusei.at[syohin, '種類'] == '株':
        zan_kabu += zan_a
    if df_zokusei.at[syohin, '種類'] == '債券':
        zan_saiken += zan_a
    if df_zokusei.at[syohin, '種類'] == 'リート':
        zan_reit += zan_a
    if df_zokusei.at[syohin, '国'] == '日本':
        zan_yen += zan_a
    if df_zokusei.at[syohin, '国'] == '米国':
        zan_usa += zan_a
    if df_zokusei.at[syohin, '国'] == '豪州':
        zan_aus += zan_a
    if df_zokusei.at[syohin, '国'] == 'ユーロ':
        zan_eur += zan_a
    if df_zokusei.at[syohin, '国'] == 'マルチ':
        zan_out += zan_a
    if df_zokusei.at[syohin, 'ＳＢ'] == 'SB':
        zan_sb += zan_a

# アウトプット
# 比率計算
rt_yokin = float(zan_yokin / zan_all * 100)
rt_teiki = float(zan_teiki / zan_all * 100)
rt_kabu = float(zan_kabu / zan_all * 100)
rt_saiken = float(zan_saiken / zan_all * 100)
rt_reit = float(zan_reit / zan_all * 100)

rt_yen = float(zan_yen / zan_all * 100)
rt_usa = float(zan_usa / zan_all * 100)
rt_eur = float(zan_eur / zan_all * 100)
rt_aus = float(zan_aus / zan_all * 100)
rt_out = float(zan_out / zan_all * 100)

rt_sb = float(zan_sb / zan_all * 100)

zan_all = '{:,}'.format(zan_all)
print('総資産ーーー＞', zan_all, '円')

print('')
print('投資商品別')
zan_yokin = '{:,}'.format(zan_yokin)
rt_yokin = '{:.1f}'.format(rt_yokin)
print('預金ーーーー＞', zan_yokin, '  ', rt_yokin, '%')
zan_teiki = '{:,}'.format(zan_teiki)
rt_teiki = '{:.1f}'.format(rt_teiki)
print('定期預金ーー＞', zan_teiki, '  ', rt_teiki, '%')
zan_kabu = '{:,}'.format(zan_kabu)
rt_kabu = '{:.1f}'.format(rt_kabu)
print('株式ーーーー＞', zan_kabu, '  ', rt_kabu, '%')
zan_saiken = '{:,}'.format(zan_saiken)
rt_saiken = '{:.1f}'.format(rt_saiken)
print('債券ーーーー＞', zan_saiken, '  ', rt_saiken, '%')
zan_reit = '{:,}'.format(zan_reit)
rt_reit = '{:.1f}'.format(rt_reit)
print('リートーーー＞', zan_reit, '  ', rt_reit, '%')

print('')
print('投資先通貨別')
zan_yen = '{:,}'.format(zan_yen)
rt_yen = '{:.1f}'.format(rt_yen)
print('円ーーーーー＞', zan_yen, '  ', rt_yen, '%')
zan_usa = '{:,}'.format(zan_usa)
rt_usa = '{:.1f}'.format(rt_usa)
print('米ーーーーー＞', zan_usa, '  ', rt_usa, '%')
zan_eur = '{:,}'.format(zan_eur)
rt_eur = '{:.1f}'.format(rt_eur)
print('ユーローーー＞', zan_eur, '  ', rt_eur, '%')
zan_aus = '{:,}'.format(zan_aus)
rt_aus = '{:.1f}'.format(rt_aus)
print('豪ーーーーー＞', zan_aus, '  ', rt_aus, '%')
zan_out = '{:,}'.format(zan_out)
rt_out = '{:.1f}'.format(rt_out)
print('マルチーーー＞', zan_out, '  ', rt_out, '%')
print('')
zan_sb = '{:,}'.format(zan_sb)
rt_sb = '{:.1f}'.format(rt_sb)
print('ソフトバンクリスク')
print('ーーーーーー＞', zan_sb, '  ', rt_sb, '%')

# EXCELへの書き込み

import openpyxl as op
wb = op.load_workbook('pandas_資産残高.xlsx')
sheet1 = wb['指標']
sheet2 = wb['推移']

sw_out = 0
cnt_retu = 1

while sw_out == 0:
    if sheet2.cell(row=1, column=cnt_retu).value == None:
        sw_out = 1
    else:
        cnt_retu += 1

sheet2.cell(row=1, column=cnt_retu).value = sheet1.cell(row=1, column=2).value
sheet2.cell(row=2, column=cnt_retu).value = sheet1.cell(row=2, column=2).value
sheet2.cell(row=3, column=cnt_retu).value = sheet1.cell(row=3, column=2).value
sheet2.cell(row=4, column=cnt_retu).value = sheet1.cell(row=4, column=2).value
sheet2.cell(row=5, column=cnt_retu).value = sheet1.cell(row=5, column=2).value
sheet2.cell(row=6, column=cnt_retu).value = sheet1.cell(row=6, column=2).value
sheet2.cell(row=7, column=cnt_retu).value = sheet1.cell(row=7, column=2).value
sheet2.cell(row=8, column=cnt_retu).value = sheet1.cell(row=8, column=2).value

sheet2.cell(row=10, column=cnt_retu).value = zan_all
sheet2.cell(row=12, column=cnt_retu).value = zan_yokin
sheet2.cell(row=13, column=cnt_retu).value = rt_yokin
sheet2.cell(row=14, column=cnt_retu).value = zan_teiki
sheet2.cell(row=15, column=cnt_retu).value = rt_teiki
sheet2.cell(row=16, column=cnt_retu).value = zan_kabu
sheet2.cell(row=17, column=cnt_retu).value = rt_kabu
sheet2.cell(row=18, column=cnt_retu).value = zan_saiken
sheet2.cell(row=19, column=cnt_retu).value = rt_saiken
sheet2.cell(row=20, column=cnt_retu).value = zan_reit
sheet2.cell(row=21, column=cnt_retu).value = rt_reit

sheet2.cell(row=23, column=cnt_retu).value = zan_yen
sheet2.cell(row=24, column=cnt_retu).value = rt_yen
sheet2.cell(row=25, column=cnt_retu).value = zan_usa
sheet2.cell(row=26, column=cnt_retu).value = rt_usa
sheet2.cell(row=27, column=cnt_retu).value = zan_eur
sheet2.cell(row=28, column=cnt_retu).value = rt_eur
sheet2.cell(row=29, column=cnt_retu).value = zan_aus
sheet2.cell(row=30, column=cnt_retu).value = rt_aus
sheet2.cell(row=31, column=cnt_retu).value = zan_out
sheet2.cell(row=32, column=cnt_retu).value = rt_out

sheet2.cell(row=34, column=cnt_retu).value = zan_sb
sheet2.cell(row=35, column=cnt_retu).value = rt_sb

wb.save('pandas_資産残高.xlsx')